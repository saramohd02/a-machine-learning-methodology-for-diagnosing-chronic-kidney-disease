from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl


# Create your views here.
from Remote_User.models import ClientRegister_Model,kidney_model,kidney_disease_model,detection_ratio_model,detection_accuracy_model


def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            kidney_model.objects.all().delete()
            kidney_disease_model.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        kidney_model.objects.create(
        id1= active_sheet.cell(r, 1).value,
        age= active_sheet.cell(r, 2).value,
        bp= active_sheet.cell(r, 3).value,
        sg= active_sheet.cell(r, 4).value,
        al= active_sheet.cell(r, 5).value,
        su= active_sheet.cell(r, 6).value,
        rbc= active_sheet.cell(r, 7).value,
        pc= active_sheet.cell(r, 8).value,
        pcc= active_sheet.cell(r, 9).value,
        ba= active_sheet.cell(r, 10).value,
        bgr= active_sheet.cell(r, 11).value,
        bu= active_sheet.cell(r, 12).value,
        sc= active_sheet.cell(r, 13).value,
        sod= active_sheet.cell(r, 14).value,
        pot= active_sheet.cell(r, 15).value,
        hemo= active_sheet.cell(r, 16).value,
        pcv= active_sheet.cell(r, 17).value,
        wc= active_sheet.cell(r, 18).value,
        rc= active_sheet.cell(r, 19).value,
        htn= active_sheet.cell(r, 20).value,
        dm= active_sheet.cell(r, 21).value,
        cad= active_sheet.cell(r, 22).value,
        appet= active_sheet.cell(r, 23).value,
        pe= active_sheet.cell(r, 24).value,
        ane= active_sheet.cell(r, 25).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})




